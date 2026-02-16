"""
Применение StyleSpec к openpyxl worksheet.

Комментарии — на русском (внешне, от третьего лица).
Print/логи — на английском (если появятся).
"""

from __future__ import annotations

from dataclasses import replace
from typing import Optional

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .models import StyleSpec, BlockStyle


def _apply_gridlines(ws: Worksheet, hide: bool) -> None:
    ws.sheet_view.showGridLines = not bool(hide)


def _apply_freeze(ws: Worksheet, freeze_rows: int, freeze_cols: int, freeze_panes: str | None) -> None:
    # Если пользователь явно передал freeze_panes, он важнее.
    if freeze_panes:
        ws.freeze_panes = freeze_panes
        return

    if freeze_rows <= 0 and freeze_cols <= 0:
        ws.freeze_panes = None
        return

    row = freeze_rows + 1 if freeze_rows > 0 else 1
    col = freeze_cols + 1 if freeze_cols > 0 else 1
    ws.freeze_panes = f"{get_column_letter(col)}{row}"


def _merge_block(base: BlockStyle, overlay: Optional[BlockStyle]) -> BlockStyle:
    """
    Накладывает overlay на base: берёт только те поля, которые не None.
    """
    if overlay is None:
        return base

    return BlockStyle(
        fill=overlay.fill if overlay.fill is not None else base.fill,
        font_color=overlay.font_color if overlay.font_color is not None else base.font_color,
        bold=overlay.bold if overlay.bold is not None else base.bold,
        border_color=overlay.border_color if overlay.border_color is not None else base.border_color,
        align_h=overlay.align_h if overlay.align_h is not None else base.align_h,
        align_v=overlay.align_v if overlay.align_v is not None else base.align_v,
        wrap_text=overlay.wrap_text if overlay.wrap_text is not None else base.wrap_text,
        number_format=overlay.number_format if overlay.number_format is not None else base.number_format,
    )


def _resolve_cell_block(spec: StyleSpec, r: int, c: int) -> BlockStyle:
    """
    Возвращает BlockStyle для конкретной ячейки по правилам 4 кирпичиков.

    По умолчанию всё равно values_block (может быть пустым).
    """
    values = spec.values_block or BlockStyle()

    in_header = spec.header_rows > 0 and r <= spec.header_rows
    in_first_cols = spec.first_cols > 0 and c <= spec.first_cols

    # corner: верх-лево
    if in_header and in_first_cols:
        return _merge_block(values, spec.corner_block or spec.header_block or spec.first_cols_block)

    # header: верх (кроме corner)
    if in_header:
        return _merge_block(values, spec.header_block)

    # first cols: лево (кроме corner)
    if in_first_cols:
        return _merge_block(values, spec.first_cols_block)

    # values: обычные значения
    return values


def _make_font(spec: StyleSpec, block: BlockStyle) -> Optional[Font]:
    """
    Формирует Font для ячейки.
    Если font_theme не задан и нет переопределений, то возвращает None (не трогать).
    """
    if spec.font_theme is None and block.font_color is None and block.bold is None:
        return None

    name = spec.font_theme.name if spec.font_theme else None
    size = spec.font_theme.size if spec.font_theme else None

    # bold: если не задано, по умолчанию False
    bold = bool(block.bold) if block.bold is not None else False

    # openpyxl: color задаётся строкой ARGB
    if block.font_color is not None:
        return Font(name=name, size=size, bold=bold, color=block.font_color)
    return Font(name=name, size=size, bold=bold)


def _make_fill(block: BlockStyle) -> Optional[PatternFill]:
    if block.fill is None:
        return None
    return PatternFill("solid", fgColor=block.fill)


def _make_alignment(block: BlockStyle) -> Optional[Alignment]:
    if block.align_h is None and block.align_v is None and block.wrap_text is None:
        return None
    return Alignment(
        horizontal=block.align_h,
        vertical=block.align_v,
        wrap_text=block.wrap_text,
    )


def _make_border(block: BlockStyle) -> Optional[Border]:
    if block.border_color is None:
        return None
    side = Side(style="thin", color=block.border_color)
    return Border(left=side, right=side, top=side, bottom=side)


def _apply_number_format(ws: Worksheet, spec: StyleSpec) -> None:
    """
    Числовой формат применяется по правилам:
    1) Если у блока ячейки задан block.number_format => применяется он.
    2) Иначе, если spec.number_decimals задан => применяется decimals-формат к числам/формулам
       (по умолчанию только вне header-строк).
    """
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    # decimals fallback формат
    fallback_fmt = None
    if spec.number_decimals is not None:
        decimals = int(spec.number_decimals)
        fallback_fmt = "0" if decimals <= 0 else "0." + ("0" * decimals)

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value

            block = _resolve_cell_block(spec, r, c)

            # 1) block-specific format
            if block.number_format:
                cell.number_format = str(block.number_format)
                continue

            # 2) fallback decimals (как раньше)
            if not fallback_fmt:
                continue

            # Не форматировать header-строки, как раньше
            if spec.header_rows > 0 and r <= spec.header_rows:
                continue

            if v is None:
                continue

            if isinstance(v, str) and v.startswith("="):
                if spec.number_apply_to_formulas:
                    cell.number_format = fallback_fmt
                continue

            if isinstance(v, (int, float)):
                cell.number_format = fallback_fmt



def apply_style(ws: Worksheet, spec: StyleSpec, *, freeze_panes: str | None = None) -> None:
    """
    Применяет StyleSpec к worksheet.
    """
    _apply_gridlines(ws, spec.hide_gridlines)
    _apply_freeze(ws, spec.freeze_rows, spec.freeze_cols, freeze_panes)

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            block = _resolve_cell_block(spec, r, c)

            font = _make_font(spec, block)
            fill = _make_fill(block)
            align = _make_alignment(block)
            border = _make_border(block)

            if font is not None:
                cell.font = font
            if fill is not None:
                cell.fill = fill
            if align is not None:
                cell.alignment = align
            if border is not None:
                cell.border = border

    _apply_number_format(ws, spec)
