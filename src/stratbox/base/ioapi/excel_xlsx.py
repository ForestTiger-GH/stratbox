"""
excel_xlsx — чтение/запись XLSX поверх FileStore.

Ключевой принцип:
- pandas пишет DataFrame в xlsx
- дальше openpyxl применяет стили/фильтр/автоширины

Автоподбор ширины спроектирован так, чтобы ориентироваться на ОТОБРАЖАЕМОЕ
значение (с учетом number_format), а не на сырое str(value).
"""

from __future__ import annotations

from io import BytesIO
from typing import Any, Optional, Tuple

import pandas as pd

from stratbox.base.filestore.base import FileStore
from stratbox.base.ioapi.bytes import read_bytes, write_bytes
from stratbox.base.utils.optional_deps import ensure_import


# -----------------------------
# Внутренние утилиты автоширины
# -----------------------------

def _safe_str(v: Any) -> str:
    """Безопасное преобразование к строке, без падений."""
    try:
        return str(v)
    except Exception:
        return ""


def _strip_and_flatten_text(s: str) -> str:
    """Убирает переносы строк и лишние пробелы."""
    s = s.replace("\r", " ").replace("\n", " ")
    return " ".join(s.split()).strip()


def _is_formula(v: Any) -> bool:
    return isinstance(v, str) and v.startswith("=")


def _parse_autofilter_columns(ws) -> Optional[Tuple[int, int]]:
    """
    Возвращает диапазон колонок (min_col, max_col), на который реально
    навешан автофильтр, или None если фильтра нет/не распарсили.

    Нужно для того, чтобы filter_padding добавлялся ТОЛЬКО тем колонкам,
    которые попали под фильтр (иначе будет лишняя ширина).
    """
    try:
        ref = getattr(ws.auto_filter, "ref", None)
        if not ref:
            return None

        # ref вида "A1:Z200" или "A1:A10"
        if ":" in ref:
            left, right = ref.split(":", 1)
        else:
            left, right = ref, ref

        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

        col_l, _ = coordinate_from_string(left)
        col_r, _ = coordinate_from_string(right)
        min_c = int(column_index_from_string(col_l))
        max_c = int(column_index_from_string(col_r))
        if min_c > max_c:
            min_c, max_c = max_c, min_c
        return (min_c, max_c)
    except Exception:
        return None


def _count_decimals_from_format(fmt: str) -> Optional[int]:
    """
    Пытается вытащить число знаков после запятой из формата вида:
    - "0.00"
    - "#,##0.000"
    - "0.0%"
    Возвращает None если не удалось.
    """
    try:
        fmt = fmt.strip()
        if not fmt:
            return None

        # убираем экранированные секции и текст в кавычках — упрощенно
        # (для автоширины достаточно грубой оценки)
        # оставим только первую секцию до ';'
        if ";" in fmt:
            fmt = fmt.split(";", 1)[0]

        # проценты: "0.0%"
        if "%" in fmt:
            fmt = fmt.replace("%", "")

        if "." not in fmt:
            return 0

        after = fmt.split(".", 1)[1]
        # количество '0' подряд после точки
        dec = 0
        for ch in after:
            if ch == "0":
                dec += 1
            else:
                break
        return dec
    except Exception:
        return None


def _looks_like_date_format(fmt: str) -> bool:
    """
    Грубая проверка "похоже ли на дату".
    """
    fmt = fmt.lower()
    # Excel date tokens: y, m, d, h, s
    # Важно: 'm' может быть и minutes, но для ширины не критично.
    has_y = "y" in fmt
    has_d = "d" in fmt
    has_m = "m" in fmt
    return (has_y and (has_m or has_d)) or (has_d and has_m)


def _format_number_for_width(value: float, fmt: str) -> str:
    """
    Превращает число в строку примерно так, как Excel будет показывать,
    чтобы оценить ширину.

    Это НЕ полноценный форматтер Excel, но для автоширины работает стабильно.
    """
    try:
        fmt0 = (fmt or "").strip()
        # берем первую секцию формата
        if ";" in fmt0:
            fmt0 = fmt0.split(";", 1)[0].strip()

        # проценты
        if "%" in fmt0:
            dec = _count_decimals_from_format(fmt0)
            dec = 0 if dec is None else int(dec)
            v = value * 100.0
            if dec <= 0:
                s = f"{v:.0f}%"
            else:
                s = f"{v:.{dec}f}%"
            return s

        # дата-форматы тут не обрабатываем — это отдельная ветка
        # научная нотация — грубо
        if "e" in fmt0.lower():
            return f"{value:.6e}"

        # десятичные
        dec = _count_decimals_from_format(fmt0)
        if dec is None:
            # общий случай
            return f"{value:.15g}"

        dec = int(dec)

        # разделители тысяч
        use_thousands = "," in fmt0

        if dec <= 0:
            if use_thousands:
                return f"{value:,.0f}"
            return f"{value:.0f}"

        if use_thousands:
            return f"{value:,.{dec}f}"
        return f"{value:.{dec}f}"

    except Exception:
        return f"{value:.15g}"


def _format_cell_for_width(cell) -> str:
    """
    Возвращает строку "примерно как отображает Excel" для оценки ширины.
    """
    v = cell.value
    if v is None:
        return ""

    # Формулы: НЕ используем текст формулы (он длинный и не соответствует видимому значению)
    # Для автоширины достаточно вернуть пусто (колонку держит min_width) или
    # можно вернуть небольшой маркер. Пусть будет пусто.
    # Формулы: openpyxl НЕ вычисляет их как Excel.
    # Но в наших выгрузках часто встречаются "формулы-литералы" вида "=12345" или "=\"text\"".
    # Их можно безопасно распарсить и использовать для оценки ширины.
    if _is_formula(v):
        f = _safe_str(v).strip()

        # 1) числовой литерал: "=123", "=-123", "=123.45"
        import re
        m_num = re.fullmatch(r"=\s*([+-]?\d+(?:\.\d+)?)\s*", f)
        if m_num:
            try:
                num = float(m_num.group(1))
                fmt = _safe_str(getattr(cell, "number_format", ""))
                return _format_number_for_width(num, fmt)
            except Exception:
                # если что-то пошло не так — просто используем текст числа
                return m_num.group(1)

        # 2) строковый литерал: ="ABC" или ="ABC DEF"
        m_txt = re.fullmatch(r'=\s*"([^"]*)"\s*', f)
        if m_txt:
            return _strip_and_flatten_text(m_txt.group(1))

        # 3) прочие формулы (SUM/VLOOKUP/…) не учитываем
        return ""


    # Текст
    if isinstance(v, str):
        return _strip_and_flatten_text(v)

    # Даты/время
    import datetime as _dt
    if isinstance(v, (_dt.date, _dt.datetime)):
        fmt = _safe_str(getattr(cell, "number_format", "")).lower()
        # если в формате есть часы/минуты — покажем время
        if "h" in fmt or "s" in fmt:
            if isinstance(v, _dt.datetime):
                return v.strftime("%Y-%m-%d %H:%M")
            # date без времени
            return _dt.datetime(v.year, v.month, v.day, 0, 0).strftime("%Y-%m-%d %H:%M")
        return v.strftime("%Y-%m-%d")

    # Числа
    if isinstance(v, (int, float)):
        fmt = _safe_str(getattr(cell, "number_format", ""))
        if fmt and _looks_like_date_format(fmt):
            # иногда дата может быть числом (Excel serial) — но тут без конвертера.
            # для ширины берём как число, но не длинно.
            return f"{float(v):.15g}"

        # формируем строку по number_format
        return _format_number_for_width(float(v), fmt)

    # Всё остальное
    return _strip_and_flatten_text(_safe_str(v))


def _fit_column_widths(
    ws,
    *,
    min_width: float | None,
    max_width: float | None,
    sample_rows: int = 2000,
    max_cell_chars: int = 60,
    include_header: bool = False,
    header_max_chars: int = 22,
    padding: float = 2.8,
    filter_padding: float = 2.0,
) -> None:
    """
    Авто-настройка ширины столбцов.

    Правила:
    - каждая колонка измеряется отдельно
    - измеряем по "отображаемым строкам" (number_format учитывается)
    - clamp по min/max обязателен
    - padding добавляется всегда
    - filter_padding добавляется только колонкам в диапазоне автофильтра
    """
    from openpyxl.utils import get_column_letter

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    last_row = min(max_row, int(sample_rows)) if sample_rows and sample_rows > 0 else max_row

    # Если заголовок не включаем — начинаем с 2 строки (если она есть)
    start_row = 1 if include_header else 2
    if start_row > last_row:
        start_row = 1

    # Какие колонки реально под автофильтром
    filter_cols = _parse_autofilter_columns(ws)

    def _col_has_filter(c: int) -> bool:
        if not filter_cols:
            return False
        return filter_cols[0] <= c <= filter_cols[1]

    for c in range(1, max_col + 1):
        try:
            best = 0  # максимальная длина отображаемой строки в колонке

            # 1) Заголовок (строка 1) — опционально и с сильным ограничением
            if include_header and max_row >= 1:
                cell = ws.cell(row=1, column=c)
                if cell.coordinate not in ws.merged_cells:
                    s = _format_cell_for_width(cell)
                    if s:
                        s = _strip_and_flatten_text(s)
                        if header_max_chars and len(s) > int(header_max_chars):
                            s = s[: int(header_max_chars)]
                        best = max(best, len(s))

            # 2) Значения
            for r in range(start_row, last_row + 1):
                cell = ws.cell(row=r, column=c)

                # merged cells: значение только в верхней-левой; остальные пропускаем
                if cell.coordinate in ws.merged_cells:
                    continue

                s = _format_cell_for_width(cell)
                if not s:
                    continue

                s = _strip_and_flatten_text(s)
                if not s:
                    continue

                if max_cell_chars and len(s) > int(max_cell_chars):
                    s = s[: int(max_cell_chars)]

                best = max(best, len(s))

            # 3) Преобразуем длину -> width units
            # Основная идея:
            # - Excel ширина чуть больше, чем длина строки в символах, из-за отступов и ширины букв.
            # - Мы добавляем стабильный коэффициент и padding.
            #
            # Важно: padding и filter_padding трактуются как "width units", а не как символы.
            if best <= 0:
                # Пустая колонка: ставим минимум (если задан), иначе не трогаем.
                if min_width is not None:
                    col_letter = get_column_letter(c)
                    ws.column_dimensions[col_letter].width = float(min_width)
                continue

            # Базовая оценка ширины. Коэффициент подобран так, чтобы ближе быть к Excel autofit.
            width = (best * 1.15) + float(padding)

            # Добавка под кнопку автофильтра только для колонок в фильтре
            if _col_has_filter(c):
                width += float(filter_padding)

            # Clamp
            if min_width is not None:
                width = max(width, float(min_width))
            if max_width is not None:
                width = min(width, float(max_width))

            col_letter = get_column_letter(c)
            ws.column_dimensions[col_letter].width = float(width)

        except Exception as e:
            print(f"[WARN] Auto width skipped column {c}: {e}")
            continue


# -----------------------------
# Публичные функции read/write
# -----------------------------

def read_df(
    path: str,
    store: FileStore | None = None,
    *,
    auto_install: bool | None = None,
    **kwargs: Any,
) -> pd.DataFrame:
    # openpyxl нужен pandas'у как engine для xlsx
    ensure_import("openpyxl", "openpyxl>=3.1", auto_install=auto_install)

    data = read_bytes(path, store=store)
    return pd.read_excel(BytesIO(data), engine="openpyxl", **kwargs)


def write_df(
    path: str,
    df: pd.DataFrame,
    store: FileStore | None = None,
    *,
    sheet_name: str = "data",
    meta: dict[str, Any] | None = None,
    style_preset: str | None = "DEFAULT",
    freeze_panes: str | None = None,
    auto_col_width: bool = True,
    col_width_min: float | None = 6.0,
    col_width_max: float | None = 60.0,
    col_width_sample_rows: int = 2000,
    col_width_include_header: bool = False,
    col_width_header_max_chars: int = 22,
    col_width_padding: float = 2.8,
    col_width_filter_padding: float = 2.0,
    auto_install: bool | None = None,
    index: bool = False,
    **kwargs: Any,
) -> None:
    """
    Пишет DataFrame в XLSX и (опционально) применяет:
    - sheet_name
    - метаданные книги (meta)
    - форматирование (style_preset)
    - автофильтр (через autofilter_range)
    - автоподбор ширины столбцов

    style_preset:
    - None => стили не применять
    - "DEFAULT" => применить дефолтный пресет (core или plugin)
    - либо конкретное имя пресета (core или plugin)

    Автофильтр:
    - можно передать autofilter_range="A1:Z200" через kwargs или engine_kwargs
    - этот параметр НЕ прокидывается в pandas, чтобы не ловить несовместимость версий
    """
    ensure_import("openpyxl", "openpyxl>=3.1", auto_install=auto_install)

    from openpyxl import load_workbook
    from stratbox.base.styles.excel.main import apply_preset

    bio = BytesIO()

    # ВАЖНО: index контролируется явно, чтобы не получить конфликт kwargs
    if "index" in kwargs:
        kwargs.pop("index")

    # ВАЖНО: pandas/openpyxl в некоторых окружениях не принимает autofilter_range в writer._write_cells.
    # Поэтому этот параметр надо извлечь и применить вручную через openpyxl уже ПОСЛЕ записи.
    autofilter_range = None

    if "autofilter_range" in kwargs:
        autofilter_range = kwargs.pop("autofilter_range")

    engine_kwargs = kwargs.pop("engine_kwargs", None)
    if isinstance(engine_kwargs, dict) and "autofilter_range" in engine_kwargs:
        autofilter_range = engine_kwargs.pop("autofilter_range")
        # Остальные engine_kwargs намеренно не прокидываются, чтобы избежать несовместимостей.

    # 1) pandas запись
    df.to_excel(bio, index=index, engine="openpyxl", sheet_name=sheet_name, **kwargs)

    # 2) пост-обработка через openpyxl
    wb = load_workbook(BytesIO(bio.getvalue()))
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    # 2.1) метаданные
    if meta:
        props = wb.properties
        if "creator" in meta:
            props.creator = _safe_str(meta["creator"])
        if "title" in meta:
            props.title = _safe_str(meta["title"])
        if "subject" in meta:
            props.subject = _safe_str(meta["subject"])
        if "category" in meta:
            props.category = _safe_str(meta["category"])
        if "keywords" in meta:
            props.keywords = _safe_str(meta["keywords"])
        if "description" in meta:
            props.description = _safe_str(meta["description"])

    # 2.2) стили (если заданы)
    if style_preset is not None:
        apply_preset(ws, style_preset, freeze_panes=freeze_panes)

    # 2.3) автофильтр — обязательно ДО автоширины
    if autofilter_range:
        try:
            ws.auto_filter.ref = _safe_str(autofilter_range)
        except Exception:
            pass

    if not autofilter_range:
        try:
            from openpyxl.utils import get_column_letter

            ncols = int(df.shape[1]) + (1 if index else 0)
            nrows = int(df.shape[0]) + 1  # +1 строка заголовка
            if ncols > 0 and nrows > 1:
                rng = f"A1:{get_column_letter(ncols)}{nrows}"
                ws.auto_filter.ref = rng
        except Exception:
            pass

    # 2.4) авто-настройка ширины столбцов
    if auto_col_width:
        try:
            _fit_column_widths(
                ws,
                min_width=col_width_min,
                max_width=col_width_max,
                sample_rows=col_width_sample_rows,
                include_header=col_width_include_header,
                header_max_chars=col_width_header_max_chars,
                padding=col_width_padding,
                filter_padding=col_width_filter_padding,
            )
        except Exception as e:
            print(f"[WARN] Auto column width failed: {e}")

    out = BytesIO()
    wb.save(out)
    write_bytes(path, out.getvalue(), store=store)
