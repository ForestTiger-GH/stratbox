"""
stratbox.base.ioapi.excel_styles

Универсальная стилизация Excel (openpyxl):
- атомарные фичи (gridlines, freeze, fonts, number formats, palettes, borders)
- пресеты, которые комбинируют фичи

Комментарии — на русском (внешне, от третьего лица).
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Optional, Dict, Tuple

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------
# Модели: шрифт и палитра
# ---------------------------

@dataclass(frozen=True)
class FontTheme:
    """Тема шрифтов."""
    name: str
    size: int
    header_bold: bool = True


@dataclass(frozen=True)
class ColorPalette:
    """
    Палитра цветов.
    Все цвета передаются в формате ARGB без '#', например: 'FF00AA00'.
    """
    name: str

    # заливки
    header_fill_main: str        # для верхних строк (заголовки)
    header_fill_side: str        # для левых столбцов (вторичные заголовки)
    header_font_color: str       # цвет текста заголовка

    # границы
    data_border: str             # цвет границ данных
    header_border: str           # цвет границ заголовков (накладывается поверх)

    # заливка обычных данных (обычно пусто/белый); можно расширить позже
    data_fill: Optional[str] = None


@dataclass
class StyleSpec:
    """
    Полная спецификация стиля.

    Принцип:
    - если параметр None => фича не применяется (не трогает существующее)
    - если число/значение => применяется
    """
    # базовое
    hide_gridlines: bool = True

    # freeze: сколько закрепить строк/столбцов (0 => не закреплять)
    freeze_rows: int = 1
    freeze_cols: int = 2

    # шрифт
    font_theme: Optional[FontTheme] = None

    # заголовки (верхние строки / левые столбцы)
    header_rows: int = 1
    header_cols: int = 2

    # числовой формат:
    # decimals=None => не применять числовой формат
    # decimals=0/1/2/3 => применять
    number_decimals: Optional[int] = 0
    number_apply_to_formulas: bool = True

    # палитра
    palette: Optional[ColorPalette] = None

    # применять границы
    apply_borders: bool = True


# ---------------------------
# Готовые темы шрифтов
# ---------------------------

FONT_THEMES: Dict[str, FontTheme] = {
    "ARIAL": FontTheme(name="Arial", size=10, header_bold=True),
    "TIMES": FontTheme(name="Times New Roman", size=12, header_bold=True),
}


# ---------------------------
# Примеры палитр (5 шт.)
# Можно потом поменять цвета, структура останется.
# ---------------------------

PALETTES: Dict[str, ColorPalette] = {
    "GREEN": ColorPalette(
        name="GREEN",
        header_fill_main="FF1F7A1F",   # тёмно-зелёный
        header_fill_side="FFE6E6E6",   # светло-серый
        header_font_color="FFFFFFFF",  # белый текст
        data_border="FFBFBFBF",        # серые границы
        header_border="FF4F4F4F",      # более тёмные границы заголовка
        data_fill=None,
    ),
    "BLUE": ColorPalette(
        name="BLUE",
        header_fill_main="FF1F4E79",
        header_fill_side="FFE6E6E6",
        header_font_color="FFFFFFFF",
        data_border="FFBFBFBF",
        header_border="FF3F3F3F",
        data_fill=None,
    ),
    "YELLOW": ColorPalette(
        name="YELLOW",
        header_fill_main="FFFFC000",
        header_fill_side="FFE6E6E6",
        header_font_color="FF000000",
        data_border="FFBFBFBF",
        header_border="FF3F3F3F",
        data_fill=None,
    ),
    "GRAY": ColorPalette(
        name="GRAY",
        header_fill_main="FF595959",
        header_fill_side="FFE6E6E6",
        header_font_color="FFFFFFFF",
        data_border="FFBFBFBF",
        header_border="FF2F2F2F",
        data_fill=None,
    ),
    "MINT": ColorPalette(
        name="MINT",
        header_fill_main="FF66CC99",
        header_fill_side="FFE6E6E6",
        header_font_color="FF000000",
        data_border="FFBFBFBF",
        header_border="FF3F3F3F",
        data_fill=None,
    ),
}


# ---------------------------
# Пресеты (комбинации)
# ---------------------------

PRESETS: Dict[str, StyleSpec] = {
    "MACROBANKS_GREEN": StyleSpec(
        hide_gridlines=True,
        freeze_rows=1,
        freeze_cols=2,
        font_theme=FONT_THEMES["ARIAL"],
        header_rows=1,
        header_cols=2,
        number_decimals=0,
        number_apply_to_formulas=True,
        palette=PALETTES["GREEN"],
        apply_borders=True,
    ),
    "MACROBANKS_BLUE": StyleSpec(
        hide_gridlines=True,
        freeze_rows=1,
        freeze_cols=2,
        font_theme=FONT_THEMES["ARIAL"],
        header_rows=1,
        header_cols=2,
        number_decimals=0,
        number_apply_to_formulas=True,
        palette=PALETTES["BLUE"],
        apply_borders=True,
    ),
}


# ---------------------------
# Вспомогательные функции
# ---------------------------

def _calc_freeze_cell(freeze_rows: int, freeze_cols: int) -> Optional[str]:
    """
    Рассчитывает ячейку freeze_panes.
    Пример: freeze_rows=1, freeze_cols=2 => 'C2'
    """
    if freeze_rows <= 0 and freeze_cols <= 0:
        return None
    row = freeze_rows + 1 if freeze_rows > 0 else 1
    col = freeze_cols + 1 if freeze_cols > 0 else 1
    return f"{get_column_letter(col)}{row}"


def _used_range(ws: Worksheet) -> Tuple[int, int]:
    """
    Возвращает (max_row, max_col) по ws.max_row/ws.max_column.
    Это быстро и достаточно для таблиц, созданных через pandas.to_excel().
    """
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    return max_row, max_col


def _make_fill(argb: str) -> PatternFill:
    return PatternFill(fill_type="solid", start_color=argb, end_color=argb)


def _make_border(argb: str) -> Border:
    side = Side(style="thin", color=argb)
    return Border(left=side, right=side, top=side, bottom=side)


def _num_format(decimals: int) -> str:
    # Excel-формат: разделители разрядов + фиксированные знаки
    if decimals <= 0:
        return "#,##0"
    return "#,##0." + ("0" * decimals)


# ---------------------------
# Главная функция применения стиля
# ---------------------------

def apply_style(ws: Worksheet, spec: StyleSpec) -> None:
    """
    Применяет style spec к листу.
    """
    max_row, max_col = _used_range(ws)

    # 1) базовое
    if spec.hide_gridlines:
        ws.sheet_view.showGridLines = False

    # 2) freeze
    freeze_cell = _calc_freeze_cell(spec.freeze_rows, spec.freeze_cols)
    if freeze_cell:
        ws.freeze_panes = freeze_cell

    # 3) шрифт
    if spec.font_theme:
        base_font = Font(name=spec.font_theme.name, size=spec.font_theme.size, bold=False)
        # общий шрифт на всё
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.font = base_font

        # заголовок жирным
        if spec.font_theme.header_bold and spec.header_rows > 0:
            header_font = Font(
                name=spec.font_theme.name,
                size=spec.font_theme.size,
                bold=True,
                color=(spec.palette.header_font_color if spec.palette else None),
            )
            for r in range(1, min(spec.header_rows, max_row) + 1):
                for c in range(1, max_col + 1):
                    ws.cell(row=r, column=c).font = header_font

    # 4) числовой формат
    if spec.number_decimals is not None:
        fmt = _num_format(spec.number_decimals)
        # Применяется к телу таблицы, начиная со 2-й строки и 3-го столбца (как у вас обычно).
        # Но пользователь может изменить header_rows/header_cols.
        start_r = spec.header_rows + 1 if spec.header_rows > 0 else 1
        start_c = spec.header_cols + 1 if spec.header_cols > 0 else 1

        for row in ws.iter_rows(min_row=start_r, max_row=max_row, min_col=start_c, max_col=max_col):
            for cell in row:
                v = cell.value
                # Числа форматируются всегда
                if isinstance(v, (int, float)):
                    cell.number_format = fmt
                # Формулы тоже форматируются, если включено
                elif spec.number_apply_to_formulas and isinstance(v, str) and v.startswith("="):
                    cell.number_format = fmt

    # 5) палитра: заливки + цвет текста заголовков
    if spec.palette:
        pal = spec.palette

        fill_main = _make_fill(pal.header_fill_main)
        fill_side = _make_fill(pal.header_fill_side)

        # верхние header_rows строк в основной цвет
        if spec.header_rows > 0:
            for r in range(1, min(spec.header_rows, max_row) + 1):
                for c in range(1, max_col + 1):
                    ws.cell(row=r, column=c).fill = fill_main

        # левые header_cols столбцов в вторичный цвет (кроме пересечения, там уже основной)
        if spec.header_cols > 0:
            for r in range(1, max_row + 1):
                for c in range(1, min(spec.header_cols, max_col) + 1):
                    if r <= spec.header_rows:
                        continue
                    ws.cell(row=r, column=c).fill = fill_side

    # 6) границы
    if spec.apply_borders and spec.palette:
        pal = spec.palette
        data_border = _make_border(pal.data_border)
        header_border = _make_border(pal.header_border)

        # Границы данных по заполненным ячейкам (по всей области таблицы)
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                # Границы ставятся только если ячейка реально используется
                if cell.value is None or cell.value == "":
                    continue
                cell.border = data_border

        # Границы заголовков поверх (верхние строки + левые столбцы)
        # Верхние строки
        if spec.header_rows > 0:
            for r in range(1, min(spec.header_rows, max_row) + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    if cell.value is None or cell.value == "":
                        continue
                    cell.border = header_border

        # Левые столбцы
        if spec.header_cols > 0:
            for r in range(1, max_row + 1):
                for c in range(1, min(spec.header_cols, max_col) + 1):
                    cell = ws.cell(row=r, column=c)
                    if cell.value is None or cell.value == "":
                        continue
                    cell.border = header_border


def resolve_style(
    *,
    preset: Optional[str] = None,
    font: Optional[str] = None,
    palette: Optional[str] = None,
    freeze_rows: Optional[int] = None,
    freeze_cols: Optional[int] = None,
    header_rows: Optional[int] = None,
    header_cols: Optional[int] = None,
    decimals: Optional[int] = None,
    number_format_enabled: Optional[bool] = None,
) -> StyleSpec:
    """
    Собирает StyleSpec:
    - берёт preset, если задан
    - затем накладывает поверх параметризацию пользователя
    """
    if preset:
        key = preset.strip().upper()
        if key not in PRESETS:
            raise ValueError(f"Unknown excel style preset: {preset}")
        spec = PRESETS[key]
        # делается “копия” dataclass через пересборку (внешне, безопасно)
        spec = StyleSpec(**spec.__dict__)
    else:
        spec = StyleSpec()

    if font:
        fkey = font.strip().upper()
        if fkey not in FONT_THEMES:
            raise ValueError(f"Unknown font theme: {font}")
        spec.font_theme = FONT_THEMES[fkey]

    if palette:
        pkey = palette.strip().upper()
        if pkey not in PALETTES:
            raise ValueError(f"Unknown palette: {palette}")
        spec.palette = PALETTES[pkey]

    if freeze_rows is not None:
        spec.freeze_rows = int(freeze_rows)
    if freeze_cols is not None:
        spec.freeze_cols = int(freeze_cols)

    if header_rows is not None:
        spec.header_rows = int(header_rows)
    if header_cols is not None:
        spec.header_cols = int(header_cols)

    if number_format_enabled is not None and not bool(number_format_enabled):
        spec.number_decimals = None

    if decimals is not None:
        spec.number_decimals = int(decimals)

    return spec

def apply_preset(ws: Worksheet, preset: str, *, freeze_panes: str | None = None) -> None:
    """
    Совместимость с ioapi/excel_xlsx.py:
    - preset: строка с именем пресета (case-insensitive)
    - freeze_panes: если задано, переопределяет freeze_panes напрямую (например "C2")
    """
    key = str(preset).strip().upper()

    # Нормализатор “старых” имён
    if key == "MACROBANKS_TABLE":
        key = "MACROBANKS_TABLE"
    elif key == "MACROBANKS_GREEN":
        key = "MACROBANKS_GREEN"
    elif key == "MACROBANKS_BLUE":
        key = "MACROBANKS_BLUE"

    if key not in PRESETS:
        raise ValueError(f"Unknown excel style preset: {preset}")

    spec = PRESETS[key]
    # Безопасная копия spec (на случай дальнейших правок)
    spec = StyleSpec(**spec.__dict__)

    # Если freeze_panes передали строкой (C2 и т.п.) — это явный приоритет пользователя.
    if freeze_panes:
        ws.freeze_panes = freeze_panes
        # Чтобы apply_style не перебил freeze_panes, отключается его расчёт через rows/cols
        spec.freeze_rows = 0
        spec.freeze_cols = 0

    apply_style(ws, spec)
