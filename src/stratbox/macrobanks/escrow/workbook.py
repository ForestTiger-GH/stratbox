"""
workbook — сборка итоговой Excel-книги по счетам эскроу.

Текущее оформление сознательно оставлено близким к исходному скрипту,
чтобы мягко перенести существующую витрину в stratbox.
"""

from __future__ import annotations

import re
from numbers import Number

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


HEADER_FONT = Font(name="Arial", size=10, bold=True, color="000000")
DATA_FONT = Font(name="Arial", size=10)
FIRST_COL_FONT = Font(name="Arial", size=10, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
PALE_GREEN_FILL = PatternFill("solid", fgColor="D8F5DC")
BORDER_STYLE = Side(border_style="thin", color="D3D3D3")
CELL_BORDER = Border(top=BORDER_STYLE, bottom=BORDER_STYLE, left=BORDER_STYLE, right=BORDER_STYLE)
WIDTH_FIRST_COL = 25
WIDTH_OTHER_COLS = 11



def abbreviate_sheet_name(text: str) -> str:
    """Строит короткое имя листа по первым буквам слов с учетом лимита Excel."""
    text_clean = re.sub(r"[^а-яА-Яa-zA-Z ]", "", str(text))
    abbr = "".join(word[0].upper() for word in text_clean.split() if word)
    return abbr[:31] or "SHEET"


def make_unique_sheet_name(base_name: str, used_names: set[str]) -> str:
    """Делает имя листа уникальным, сохраняя ограничение Excel в 31 символ."""
    if base_name not in used_names:
        used_names.add(base_name)
        return base_name

    counter = 2
    while True:
        suffix = f"_{counter}"
        candidate = f"{base_name[: 31 - len(suffix)]}{suffix}"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        counter += 1



def build_escrow_workbook(
    pivots: dict[str, pd.DataFrame],
    indicators_order: list[str],
    *,
    title_row_font: Font | None = None,
    show_progress: bool = True,
) -> Workbook:
    """Формирует книгу Excel: один показатель — один лист."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    title_font = title_row_font or Font(name="Times New Roman", size=10, bold=True)
    iterator = indicators_order
    if show_progress:
        try:
            from tqdm.auto import tqdm

            iterator = tqdm(indicators_order, desc="Escrow workbook", leave=False)
        except Exception:
            iterator = indicators_order

    used_sheet_names: set[str] = set()

    for indicator in iterator:
        pivot_df = pivots[indicator].reset_index()
        sheet_name = make_unique_sheet_name(abbreviate_sheet_name(indicator), used_sheet_names)
        worksheet = workbook.create_sheet(title=sheet_name)

        worksheet["A2"] = str(indicator).upper()
        worksheet["A2"].font = title_font

        rows = list(dataframe_to_rows(pivot_df, index=False, header=True))
        start_row = 4
        for row_idx, row in enumerate(rows, start_row):
            highlight_row = False
            if row_idx > start_row:
                region_value = str(row[0])
                if " ФО" in region_value or "Итого по РФ" in region_value:
                    highlight_row = True

            for col_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)

                if row_idx == start_row:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.font = DATA_FONT
                    cell.alignment = Alignment(
                        horizontal="right" if col_idx > 1 else "left",
                        vertical="center",
                    )
                    if col_idx == 1:
                        cell.font = FIRST_COL_FONT
                    if col_idx > 1 and pd.notnull(value):
                        if isinstance(value, Number):
                            cell.number_format = "#,##0"
                    if highlight_row:
                        cell.fill = PALE_GREEN_FILL

                cell.border = CELL_BORDER

        for col_idx in range(1, worksheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = (
                WIDTH_FIRST_COL if col_idx == 1 else WIDTH_OTHER_COLS
            )

        worksheet.freeze_panes = "B5"
        worksheet.sheet_view.showGridLines = False

    return workbook


__all__ = ["abbreviate_sheet_name", "build_escrow_workbook", "make_unique_sheet_name"]
